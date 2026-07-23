import os
import re
import time
import threading
from datetime import datetime

import serial
from flask import Flask, jsonify, request, send_file, render_template_string
from openpyxl import Workbook, load_workbook


# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================

SERIAL_PORT = "COM3"      # Cambiar si tu Jetson usa otro puerto
BAUD_RATE = 115200

EXCEL_PATH = "Anexo_1_datos_iot.xlsx"
SHEET_NAME = "Datos extraidos"

# Datos experimentales que NO vienen en la trama.
# Cámbialos según el bloque de medición.
LOMBRICES_AGREGADAS_G = 50
PESO_CAMA_5CM_G = 1222
PESO_CAMA_10CM_G = 1610
PESO_LOMBRICES_INYECTADAS_G = 50

# Si tu Arduino/ESP32 solo manda la línea TRAMA, déjalo en False.
# Si manda TRAMA + bloque detallado, déjalo en True.
ESPERAR_BLOQUE_DETALLADO = True

excel_lock = threading.Lock()
ultimo_dato = None
app = Flask(__name__)


# ============================================================
# COLUMNAS DEL ANEXO 1
# ============================================================

HEADERS = [
    "lombrices_agregadas_g",
    "replica",
    "peso_cama_5cm_g",
    "peso_cama_10cm_g",
    "peso_lombrices_inyectadas_g",
    "trama_id",
    "estado",
    "humedad_pct",
    "temperatura_c",
    "uv_intensity",
    "conductividad_us_cm",
    "ph",
    "nitrogeno_mg_kg",
    "fosforo_mg_kg",
    "potasio_mg_kg",
    "acel_x_g",
    "acel_y_g",
    "acel_z_g",
    "lluvia_binario",
    "salud",
    "trama_raw",
    "humedad_trama_pct",
    "temperatura_trama_c",
    "conductividad_trama_us_cm",
    "ph_trama",
    "nitrogeno_trama_mg_kg",
    "fosforo_trama_mg_kg",
    "potasio_trama_mg_kg",
]


# ============================================================
# FUNCIONES DE EXCEL
# ============================================================

def inicializar_excel():
    """
    Crea el Excel si no existe.
    Si ya existe, no lo borra.
    """
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)
        print(f"[OK] Excel creado: {EXCEL_PATH}")
    else:
        print(f"[OK] Excel existente: {EXCEL_PATH}")


def obtener_siguiente_replica(ws):
    """
    Calcula la siguiente réplica para el mismo peso de lombrices.
    """
    max_replica = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        lombrices = row[0]
        replica = row[1]

        if lombrices == LOMBRICES_AGREGADAS_G and isinstance(replica, int):
            max_replica = max(max_replica, replica)

    return max_replica + 1


def guardar_en_excel(dato):
    """
    Guarda una fila en el Excel.
    """
    global ultimo_dato

    with excel_lock:
        inicializar_excel()

        wb = load_workbook(EXCEL_PATH)

        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(HEADERS)
        else:
            ws = wb[SHEET_NAME]

        replica = obtener_siguiente_replica(ws)

        fila = [
            LOMBRICES_AGREGADAS_G,
            replica,
            PESO_CAMA_5CM_G,
            PESO_CAMA_10CM_G,
            PESO_LOMBRICES_INYECTADAS_G,
            dato.get("trama_id"),
            dato.get("estado"),
            dato.get("humedad_pct"),
            dato.get("temperatura_c"),
            dato.get("uv_intensity"),
            dato.get("conductividad_us_cm"),
            dato.get("ph"),
            dato.get("nitrogeno_mg_kg"),
            dato.get("fosforo_mg_kg"),
            dato.get("potasio_mg_kg"),
            dato.get("acel_x_g"),
            dato.get("acel_y_g"),
            dato.get("acel_z_g"),
            dato.get("lluvia_binario"),
            dato.get("salud"),
            dato.get("trama_raw"),
            dato.get("humedad_trama_pct"),
            dato.get("temperatura_trama_c"),
            dato.get("conductividad_trama_us_cm"),
            dato.get("ph_trama"),
            dato.get("nitrogeno_trama_mg_kg"),
            dato.get("fosforo_trama_mg_kg"),
            dato.get("potasio_trama_mg_kg"),
        ]

        ws.append(fila)
        wb.save(EXCEL_PATH)
        wb.close()

        ultimo_dato = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "replica": replica,
            **dato
        }

        print(f"[GUARDADO] Replica {replica} | Trama {dato.get('trama_id')} | Estado {dato.get('estado')}")


# ============================================================
# PROCESAMIENTO DE TRAMA
# ============================================================

def limpiar_numero(texto):
    """
    Extrae un número de una línea de texto.
    Ejemplo: 'Humedad: 100.00 %' -> 100.00
    """
    if texto is None:
        return None

    match = re.search(r"[-+]?\d*\.?\d+", texto)
    if match:
        valor = match.group(0)
        if "." in valor:
            return float(valor)
        return int(valor)

    return None


def parsear_trama(linea):
    """
    Procesa una línea tipo:

    TRAMA: 33897,1,0.58,0.06,0.88,34.31,ATENCION,100.00,27.40,17.54,6118,7.20,1238,1999,1999;
    """

    linea_original = linea.strip()

    if "TRAMA:" not in linea_original:
        raise ValueError("La línea no contiene TRAMA")

    contenido = linea_original.replace("TRAMA:", "").replace(";", "").strip()
    partes = [p.strip() for p in contenido.split(",")]

    if len(partes) < 15:
        raise ValueError(f"Trama incompleta. Campos recibidos: {len(partes)}")

    trama_id = int(partes[0])
    lluvia_binario = int(float(partes[1]))
    salud = float(partes[2])
    acel_x = float(partes[3])
    acel_y = float(partes[4])
    acel_z = float(partes[5])
    estado = partes[6]
    humedad = float(partes[7])
    temperatura = float(partes[8])

    # Este campo en la trama parece ser un valor crudo/intermedio de UV.
    uv_trama = float(partes[9])

    conductividad = int(float(partes[10]))
    ph = float(partes[11])
    nitrogeno = int(float(partes[12]))
    fosforo = int(float(partes[13]))
    potasio = int(float(partes[14]))

    dato = {
        "trama_id": trama_id,
        "estado": estado,
        "humedad_pct": humedad,
        "temperatura_c": temperatura,

        # Si no llega el bloque detallado, se deja temporalmente el valor crudo.
        # Si luego llega 'UV_intensity: 207', se reemplaza por 207.
        "uv_intensity": uv_trama,

        "conductividad_us_cm": conductividad,
        "ph": ph,
        "nitrogeno_mg_kg": nitrogeno,
        "fosforo_mg_kg": fosforo,
        "potasio_mg_kg": potasio,
        "acel_x_g": acel_x,
        "acel_y_g": acel_y,
        "acel_z_g": acel_z,
        "lluvia_binario": lluvia_binario,
        "salud": salud,
        "trama_raw": linea_original,

        # Auditoría de valores que vienen directamente de la trama
        "humedad_trama_pct": humedad,
        "temperatura_trama_c": temperatura,
        "conductividad_trama_us_cm": conductividad,
        "ph_trama": ph,
        "nitrogeno_trama_mg_kg": nitrogeno,
        "fosforo_trama_mg_kg": fosforo,
        "potasio_trama_mg_kg": potasio,
    }

    return dato


def actualizar_con_linea_detallada(dato, linea):
    """
    Actualiza el dato usando líneas como:
    Humedad:      100.00 %
    Temperatura:  27.40 °C
    UV_intensity:  207 _
    """

    linea = linea.strip()

    if linea.startswith("Humedad:"):
        dato["humedad_pct"] = limpiar_numero(linea)

    elif linea.startswith("Temperatura:"):
        dato["temperatura_c"] = limpiar_numero(linea)

    elif linea.startswith("UV_intensity:"):
        dato["uv_intensity"] = limpiar_numero(linea)

    elif linea.startswith("Cond. Elec.:"):
        dato["conductividad_us_cm"] = limpiar_numero(linea)

    elif linea.startswith("pH:"):
        dato["ph"] = limpiar_numero(linea)

    elif linea.startswith("Nitrógeno"):
        dato["nitrogeno_mg_kg"] = limpiar_numero(linea)

    elif linea.startswith("Fósforo"):
        dato["fosforo_mg_kg"] = limpiar_numero(linea)

    elif linea.startswith("Potasio"):
        dato["potasio_mg_kg"] = limpiar_numero(linea)

    elif linea.startswith("acelerometro X:"):
        dato["acel_x_g"] = limpiar_numero(linea)

    elif linea.startswith("acelerometro Y:"):
        dato["acel_y_g"] = limpiar_numero(linea)

    elif linea.startswith("acelerometro Z:"):
        dato["acel_z_g"] = limpiar_numero(linea)

    elif linea.startswith("Lluvia binario:"):
        dato["lluvia_binario"] = limpiar_numero(linea)

    elif linea.startswith("Salud:"):
        dato["salud"] = limpiar_numero(linea)

    return dato


# ============================================================
# HILO INFINITO DE LECTURA SERIAL
# ============================================================

def lector_serial_infinito():
    """
    Lee infinitamente el puerto serial.
    Cada vez que recibe una trama completa, la guarda en Excel.
    """

    dato_actual = None

    while True:
        try:
            print(f"[SERIAL] Conectando a {SERIAL_PORT} @ {BAUD_RATE}...")

            with serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=1) as ser:
                print("[SERIAL] Conectado. Leyendo datos...")

                while True:
                    raw = ser.readline()

                    if not raw:
                        continue

                    linea = raw.decode("utf-8", errors="ignore").strip()

                    if not linea:
                        continue

                    print(f"[RX] {linea}")

                    # Caso 1: llega una nueva TRAMA
                    if linea.startswith("TRAMA:"):
                        # Si había una trama anterior sin guardar, la guardamos antes de iniciar otra.
                        if dato_actual is not None:
                            guardar_en_excel(dato_actual)
                            dato_actual = None

                        try:
                            dato_actual = parsear_trama(linea)

                            if not ESPERAR_BLOQUE_DETALLADO:
                                guardar_en_excel(dato_actual)
                                dato_actual = None

                        except Exception as e:
                            print(f"[ERROR] No se pudo procesar la trama: {e}")
                            dato_actual = None

                    # Caso 2: llegan líneas detalladas después de la TRAMA
                    elif dato_actual is not None:
                        dato_actual = actualizar_con_linea_detallada(dato_actual, linea)

                        # En tu bloque, Salud es la última línea.
                        # Cuando llega Salud, se considera lectura completa.
                        if linea.startswith("Salud:"):
                            guardar_en_excel(dato_actual)
                            dato_actual = None

        except serial.SerialException as e:
            print(f"[SERIAL ERROR] {e}")
            print("[SERIAL] Reintentando conexión en 5 segundos...")
            time.sleep(5)

        except Exception as e:
            print(f"[ERROR GENERAL] {e}")
            time.sleep(5)


# ============================================================
# ENDPOINTS FLASK
# ============================================================

@app.route("/")
def index():
    """
    Vista sencilla tipo tabla para revisar los últimos datos del Excel.
    """

    inicializar_excel()

    with excel_lock:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb[SHEET_NAME]

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

    headers = rows[0] if rows else HEADERS
    data = rows[-20:] if len(rows) > 1 else []

    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Jetson Nano IoT Logger</title>
        <meta charset="utf-8">
        <meta http-equiv="refresh" content="5">
        <style>
            body {
                font-family: Arial, sans-serif;
                margin: 20px;
                background: #f5f5f5;
            }
            h1 {
                color: #222;
            }
            table {
                border-collapse: collapse;
                width: 100%;
                background: white;
                font-size: 12px;
            }
            th, td {
                border: 1px solid #ccc;
                padding: 6px;
                text-align: center;
            }
            th {
                background: #222;
                color: white;
                position: sticky;
                top: 0;
            }
            .btn {
                display: inline-block;
                padding: 10px 14px;
                background: #0078D4;
                color: white;
                text-decoration: none;
                border-radius: 5px;
                margin-bottom: 15px;
            }
            .status {
                padding: 10px;
                background: white;
                margin-bottom: 15px;
                border-left: 5px solid #0078D4;
            }
        </style>
    </head>
    <body>
        <h1>Jetson Nano IoT Logger</h1>

        <div class="status">
            <b>Excel:</b> {{ excel_path }}<br>
            <b>Último dato:</b> {{ ultimo_dato }}
        </div>

        <a class="btn" href="/download">Descargar Excel</a>

        <h2>Últimas 20 lecturas</h2>

        <table>
            <thead>
                <tr>
                    {% for h in headers %}
                    <th>{{ h }}</th>
                    {% endfor %}
                </tr>
            </thead>
            <tbody>
                {% for row in data %}
                <tr>
                    {% for value in row %}
                    <td>{{ value }}</td>
                    {% endfor %}
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </body>
    </html>
    """

    return render_template_string(
        html,
        headers=headers,
        data=data[1:] if data and data[0] == headers else data,
        excel_path=EXCEL_PATH,
        ultimo_dato=ultimo_dato
    )


@app.route("/download")
def download_excel():
    """
    Descarga el archivo Excel.
    """
    inicializar_excel()
    return send_file(EXCEL_PATH, as_attachment=True)


@app.route("/status")
def status():
    """
    Estado del sistema.
    """
    return jsonify({
        "status": "ok",
        "excel_path": EXCEL_PATH,
        "sheet_name": SHEET_NAME,
        "ultimo_dato": ultimo_dato,
        "serial_port": SERIAL_PORT,
        "baud_rate": BAUD_RATE
    })


@app.route("/api/trama", methods=["POST"])
def api_trama():
    """
    Endpoint para probar sin serial.
    Puedes enviar una trama por HTTP.
    """

    data = request.get_json(force=True)
    trama = data.get("trama")

    if not trama:
        return jsonify({"error": "Debes enviar el campo 'trama'"}), 400

    try:
        dato = parsear_trama(trama)
        guardar_en_excel(dato)
        return jsonify({
            "status": "guardado",
            "dato": dato
        })

    except Exception as e:
        return jsonify({
            "error": str(e)
        }), 400


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    inicializar_excel()

    hilo_serial = threading.Thread(target=lector_serial_infinito, daemon=True)
    hilo_serial.start()

    app.run(
        host="0.0.0.0",
        port=5000,
        debug=False
    )